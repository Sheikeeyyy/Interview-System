import os, uuid, json, re, sqlite3, base64, csv
from datetime import datetime

from flask import Flask, render_template, request, redirect, session, flash, jsonify, url_for, send_from_directory
from flask_wtf import FlaskForm, CSRFProtect
from wtforms import FileField, IntegerField, StringField, PasswordField, SubmitField, SelectField
from wtforms.validators import DataRequired, NumberRange
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
load_dotenv()

# AI + Docs
import pdfplumber, docx
from langchain_groq import ChatGroq
from langchain_core.prompts import ChatPromptTemplate

# Whisper
import whisper
import torch
import random
import smtplib
from email.message import EmailMessage
from openpyxl import Workbook

# ---------------- APP CONFIG ----------------
app = Flask(__name__)
app.secret_key = os.getenv("SECRET_KEY", "secret")
csrf = CSRFProtect(app)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

UPLOAD_FOLDER = os.path.join(BASE_DIR, "uploads")
AUDIO_FOLDER = os.path.join(BASE_DIR, "audio")
VIDEO_FOLDER = os.path.join(BASE_DIR, "videos")
DB = os.path.join(BASE_DIR, "database.db")

# ✅ Proctoring folders
PROCTOR_SCREENSHOT_FOLDER = os.path.join(BASE_DIR, "proctor_screenshots")
PROCTOR_LOG_FOLDER = os.path.join(BASE_DIR, "proctor_logs")
PROCTOR_LOG_FILE = os.path.join(PROCTOR_LOG_FOLDER, "warnings_log.csv")

os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(AUDIO_FOLDER, exist_ok=True)
os.makedirs(VIDEO_FOLDER, exist_ok=True)
os.makedirs(PROCTOR_SCREENSHOT_FOLDER, exist_ok=True)
os.makedirs(PROCTOR_LOG_FOLDER, exist_ok=True)

print("✅ VIDEO SAVE PATH:", VIDEO_FOLDER)
print("✅ SCREENSHOT SAVE PATH:", PROCTOR_SCREENSHOT_FOLDER)
print("✅ WARNING LOG FILE:", PROCTOR_LOG_FILE)

ADMIN_USERNAME = "admin"
ADMIN_PASSWORD = "admin123"

# ---------------- LOAD WHISPER ----------------
print("Loading Whisper model...")
whisper_model = whisper.load_model("small")

# ---------------- DATABASE ----------------
def get_db():
    con = sqlite3.connect(DB)
    con.row_factory = sqlite3.Row
    return con
def ensure_candidate_email_column():
    with get_db() as con:
        cols = con.execute("PRAGMA table_info(submissions)").fetchall()
        names = [c["name"] for c in cols]
        if "candidate_email" not in names:
            con.execute(
                "ALTER TABLE submissions ADD COLUMN candidate_email TEXT"
            )



with get_db() as con:
    con.execute("""
    CREATE TABLE IF NOT EXISTS tests (
        id TEXT PRIMARY KEY,
        title TEXT,
        questions TEXT,
        answers TEXT,
        duration INTEGER,
        exam_type TEXT
    )
    """)
    con.execute("""
    CREATE TABLE IF NOT EXISTS submissions (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        test_id TEXT,
        candidate TEXT,
        score INTEGER,
        total INTEGER,
        answers TEXT,
        transcript TEXT,
        audio_files TEXT,
        ai_feedback TEXT,
        evaluated INTEGER DEFAULT 0
    )
    """)
ensure_candidate_email_column()
# ✅ Safe migration add video_file
def ensure_video_column():
    with get_db() as con:
        cols = con.execute("PRAGMA table_info(submissions)").fetchall()
        colnames = [c["name"] for c in cols]
        if "video_file" not in colnames:
            con.execute("ALTER TABLE submissions ADD COLUMN video_file TEXT")

ensure_video_column()

# ✅ Webcam per-question videos table
def ensure_webcam_tables():
    with get_db() as con:
        con.execute("""
        CREATE TABLE IF NOT EXISTS webcam_answers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            submission_id INTEGER,
            test_id TEXT,
            candidate TEXT,
            question_index INTEGER,
            question_text TEXT,
            video_file TEXT,
            created_at TEXT
        )
        """)

ensure_webcam_tables()

# ✅ Ensure transcript column exists
def ensure_webcam_transcript_column():
    with get_db() as con:
        cols = con.execute("PRAGMA table_info(webcam_answers)").fetchall()
        names = [c["name"] for c in cols]
        if "transcript" not in names:
            con.execute("ALTER TABLE webcam_answers ADD COLUMN transcript TEXT")
            print("✅ Added 'transcript' column to webcam_answers")

ensure_webcam_transcript_column()

# ---------------- AI ----------------
llm = ChatGroq(
    api_key=os.getenv("GROQ_API_KEY"),
    model="llama-3.3-70b-versatile",
    temperature=0
)

mcq_prompt = ChatPromptTemplate.from_template("""
Generate exactly {n} MCQs.

Text: {context}

Return JSON:
[
  {{"question":"Q","options":["A)","B)","C)","D)"],"answer":"A"}}
]
""")

desc_prompt = ChatPromptTemplate.from_template("""
Generate exactly {n} descriptive questions.

Text: {context}

Return JSON:
[
  {{"question":"Explain ..."}}
]
""")

eval_prompt = ChatPromptTemplate.from_template("""
You are an examiner.

Question:
{question}

Student Answer:
{answer}

Evaluate out of {max_marks}.

Return JSON:
{{
  "score": number,
  "reason": "short reason"
}}
""")

mcq_chain = mcq_prompt | llm
desc_chain = desc_prompt | llm
eval_chain = eval_prompt | llm

# ---------------- HELPERS ----------------
def extract_text(path):
    if path.endswith(".pdf"):
        with pdfplumber.open(path) as pdf:
            return "\n".join(p.extract_text() or "" for p in pdf.pages)
    if path.endswith(".docx"):
        d = docx.Document(path)
        return "\n".join(p.text for p in d.paragraphs)
    return open(path, encoding="utf-8").read()

def safe_json(raw):
    return json.loads(re.search(r"\[.*\]", raw, re.S).group())

def transcribe_audio(path):
    audio = whisper.load_audio(path)
    audio = audio[audio.nonzero()[0][0]:]  
    audio = whisper.pad_or_trim(audio)
    mel = whisper.log_mel_spectrogram(audio).to(whisper_model.device)
    options = whisper.DecodingOptions(
        fp16=torch.cuda.is_available(),
        language="en"
    )

    result = whisper.decode(whisper_model, mel, options)
    return result.text.strip()
def extract_audio_from_video(video_path, audio_path):
    os.system(
        f'ffmpeg -y -i "{video_path}" -vn -acodec pcm_s16le -ar 16000 -ac 1 "{audio_path}"'
    )


def ai_evaluate(questions, answers, max_marks=5):
    feedback = {}
    total = 0
    for i, q in enumerate(questions):
        raw = eval_chain.invoke({
            "question": q["question"],
            "answer": answers.get(str(i), ""),
            "max_marks": max_marks
        }).content
        try:
            data = json.loads(re.search(r"\{.*\}", raw, re.S).group())
        except:
            data = {"score": 0, "reason": "Evaluation error"}
        feedback[str(i)] = data
        total += int(data["score"])
    return feedback, total
def send_otp(email, otp):
    msg = EmailMessage()
    msg.set_content(f"Your OTP for the exam is: {otp}")
    msg["Subject"] = "Exam OTP Verification"
    msg["From"] = os.getenv("SMTP_EMAIL")
    msg["To"] = email

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
        smtp.login(
            os.getenv("SMTP_EMAIL"),
            os.getenv("SMTP_PASSWORD")
        )
        smtp.send_message(msg)


# ---------------- FORMS ----------------
class LoginForm(FlaskForm):
    username = StringField("Username", validators=[DataRequired()])
    password = PasswordField("Password", validators=[DataRequired()])
    submit = SubmitField("Login")

class CreateTestForm(FlaskForm):
    file = FileField("Document (Optional)")
    num = IntegerField("Questions", default=5, validators=[NumberRange(1, 50)])
    duration = IntegerField("Duration (min)", default=10)
    title = StringField("Test Title")

    exam_type = SelectField("Exam Type", choices=[
        ("mcq", "MCQ"),
        ("descriptive", "Descriptive"),
        ("voice", "Voice Interview"),
        ("webcam", "Webcam Interview")
    ])

    submit = SubmitField("Create Test")

# ---------------- ROUTES ----------------
@app.route("/login", methods=["GET", "POST"])
def login():
    form = LoginForm()
    if form.validate_on_submit():
        if form.username.data == ADMIN_USERNAME and form.password.data == ADMIN_PASSWORD:
            session["admin"] = True
            return redirect("/")
        flash("Invalid credentials", "danger")
    return render_template("login.html", form=form)

@app.route("/")
def admin():
    if not session.get("admin"):
        return redirect("/login")

    form = CreateTestForm()
    with get_db() as con:
        tests_raw = con.execute("SELECT * FROM tests").fetchall()
        
        # Parse questions JSON and add count
        tests = []
        for t in tests_raw:
            test_dict = dict(t)
            # Parse questions JSON to get actual count
            questions = json.loads(test_dict['questions'] or '[]')
            test_dict['question_count'] = len(questions)
            tests.append(test_dict)
        
        results = con.execute("""
            SELECT s.*, t.title, t.exam_type
            FROM submissions s
            JOIN tests t ON s.test_id = t.id
            ORDER BY s.id DESC
        """).fetchall()

    return render_template(
        "index.html",
        form=form,
        tests=tests,
        results=results
    )

@app.route("/create", methods=["POST"])
def create():
    form = CreateTestForm()
    if not form.validate_on_submit():
        return redirect("/")

    file = request.files.get("file")
    if file and file.filename:
        filename = secure_filename(file.filename)
        path = os.path.join(UPLOAD_FOLDER, filename)
        file.save(path)
        context = extract_text(path)
    else:
        context = "General interview and knowledge questions."

    if form.exam_type.data == "webcam":
        raw = desc_chain.invoke({"context": context, "n": form.num.data}).content
        data = safe_json(raw)
        questions = [{"question": q["question"]} for q in data]
        answers = []

    elif form.exam_type.data == "mcq":
        raw = mcq_chain.invoke({"context": context, "n": form.num.data}).content
        data = safe_json(raw)
        questions = [{"question": q["question"], "options": q["options"]} for q in data]
        answers = [q["answer"] for q in data]

    else:
        raw = desc_chain.invoke({"context": context, "n": form.num.data}).content
        data = safe_json(raw)
        questions = [{"question": q["question"]} for q in data]
        answers = []

    test_id = uuid.uuid4().hex[:8]
    with get_db() as con:
        con.execute("""
        INSERT INTO tests VALUES (?,?,?,?,?,?)
        """, (test_id, form.title.data or f"Test {test_id}",
              json.dumps(questions), json.dumps(answers),
              form.duration.data, form.exam_type.data))

    flash("Test created", "success")
    return redirect("/")

@app.route("/exam/<test_id>", methods=["GET", "POST"])
@csrf.exempt
def exam(test_id):
    # ---------------- GET TEST ----------------
    with get_db() as con:
        test = con.execute(
            "SELECT * FROM tests WHERE id=?",
            (test_id,)
        ).fetchone()

    if not test:
        return "Invalid test link", 404

    questions = json.loads(test["questions"])
    answers = json.loads(test["answers"] or "[]")
    total = len(questions)

    # =================================================
    # STEP 1: SEND OTP (NAME + EMAIL)
    # =================================================
    if request.method == "POST" and request.form.get("step") == "send_otp":

        candidate = request.form.get("candidate", "").strip()
        email = request.form.get("email", "").strip().lower()

        if not candidate or not email:
            return "Name and Email required", 400

        # 🚫 BLOCK RE-ATTEMPT BY EMAIL (BEFORE EXAM)
        with get_db() as con:
            exists = con.execute("""
                SELECT 1 FROM submissions
                WHERE test_id=? AND candidate_email=?
            """, (test_id, candidate)).fetchone()

        if exists:
            return "You have already submitted this exam.", 403

        otp = str(random.randint(100000, 999999))

        session["otp"] = otp
        session["candidate"] = candidate
        session["otp_email"] = email   
        session["verified"] = False

        send_otp(email, otp)

        return render_template(
            "exam.html",
            title=test["title"],
            step="verify_otp"
        )

    # =================================================
    # STEP 2: VERIFY OTP
    # =================================================
    if request.method == "POST" and request.form.get("step") == "verify_otp":

        user_otp = request.form.get("otp", "").strip()

        if user_otp != session.get("otp"):
            return "Invalid OTP", 403

        session["verified"] = True

        # ✅ Webcam exams go to webcam page
        if test["exam_type"] == "webcam":
            return redirect(url_for("webcam_interview", test_id=test_id))

        return render_template(
            "exam.html",
            title=test["title"],
            step="exam",
            questions=enumerate(questions),
            duration=test["duration"],
            total=total,
            exam_type=test["exam_type"]
        )

    # =================================================
    # STEP 3: SUBMIT EXAM (AFTER OTP VERIFIED)
    # =================================================
    if request.method == "POST" and session.get("verified"):

        candidate = session.get("candidate")

        # ---------- MCQ ----------
        if test["exam_type"] == "mcq":
            score = sum(
                request.form.get(str(i)) == answers[i]
                for i in range(total)
            )
            saved = None
            transcript = None
            audio_files = None

        # ---------- DESCRIPTIVE ----------
        elif test["exam_type"] == "descriptive":
            score = None
            saved = json.dumps({
                str(i): request.form.get(f"desc_{i}")
                for i in range(total)
            })
            transcript = None
            audio_files = None

# ---------- VOICE ----------
        elif test["exam_type"] == "voice":
            score = None
            saved = None

            audio_map = json.loads(request.form.get("audio_files_json", "{}"))
            transcripts = {}
            for qindex, audio_file in audio_map.items():
                path = os.path.join(AUDIO_FOLDER, audio_file)
                text = transcribe_audio(path)
                transcripts[qindex] = text if text else "[No clear speech detected]"

            transcript = json.dumps(transcripts)
            audio_files = json.dumps(audio_map)


        with get_db() as con:
            con.execute("""
                INSERT INTO submissions
                (test_id, candidate, candidate_email, score, total,
                answers, transcript, audio_files)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                test_id,
                candidate,
                session.get("otp_email"),
                score,
                total,
                saved,
                transcript,
                audio_files
            ))

        # ⏱ AUTO LOGOUT
        session.clear()

        return render_template(
            "result.html",
            candidate=candidate
        )

    # =================================================
    # DEFAULT: ASK NAME + EMAIL
    # =================================================
    return render_template(
        "exam.html",
        title=test["title"],
        step="email"
    )
    # ---------------- GET : SHOW EXAM ----------------
   

# ✅ Webcam Interview Page
@app.route("/webcam/<test_id>")
def webcam_interview(test_id):

    # ✅ require OTP verification first
    if not session.get("verified"):
        return redirect(url_for("exam", test_id=test_id))

    with get_db() as con:
        test = con.execute(
            "SELECT * FROM tests WHERE id=?",
            (test_id,)
        ).fetchone()

    questions = json.loads(test["questions"] or "[]")
    return render_template(
        "webcam.html",
        test_id=test_id,
        title=test["title"],
        questions=questions
    )

# ✅ Upload per question video
@app.route("/upload_video_question/<test_id>", methods=["POST"])
@csrf.exempt
def upload_video_question(test_id):
    if not session.get("verified"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    candidate = session.get("candidate")
    submission_id = request.form.get("submission_id")
    question_index = request.form.get("question_index")
    question_text = request.form.get("question_text")
    f = request.files.get("video")

    if not f:
        return jsonify({"ok": False, "error": "Video missing"}), 400

    # 1️⃣ Create submission if needed
    if not submission_id:
        with get_db() as con:
            con.execute("""
                INSERT INTO submissions (test_id, candidate, candidate_email)
                VALUES (?, ?, ?)
            """, (test_id, candidate, session.get("otp_email")))
            submission_id = con.execute(
                "SELECT last_insert_rowid()"
            ).fetchone()[0]
    else:
        submission_id = int(submission_id)

    # 2️⃣ Save video
    video_filename = f"{uuid.uuid4().hex}.webm"
    video_path = os.path.join(VIDEO_FOLDER, video_filename)
    f.save(video_path)

    print("✅ Saved question video:", video_path)

    # 3️⃣ Extract audio
    audio_filename = video_filename.replace(".webm", ".wav")
    audio_path = os.path.join(AUDIO_FOLDER, audio_filename)

    extract_audio_from_video(video_path, audio_path)

    # 4️⃣ Transcribe
    try:
        transcript_text = transcribe_audio(audio_path)
        if not transcript_text:
            transcript_text = "[No clear speech detected]"
    except Exception as e:
        transcript_text = "[Transcription failed]"

    # 5️⃣ Save to DB
    with get_db() as con:
        con.execute("""
            INSERT INTO webcam_answers
            (submission_id, test_id, candidate,
             question_index, question_text,
             video_file, transcript, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        """, (
            submission_id,
            test_id,
            candidate,
            int(question_index),
            question_text,
            video_filename,
            transcript_text,
            datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))

    # 6️⃣ RETURN ONCE
    return jsonify({
        "ok": True,
        "submission_id": submission_id,
        "message": "NEXT_QUESTION"
    })

# ✅ Warning + screenshot save route
@app.route("/proctor_warning/<test_id>", methods=["POST"])
@csrf.exempt

def proctor_warning(test_id):
    print("🔥 PROCTOR WARNING ROUTE HIT")

    
    data = request.get_json(force=True)

    candidate = data.get("candidate", "unknown")
    reason = data.get("reason", "UNKNOWN")
    face_count = data.get("face_count", 0)
    gaze_x = data.get("gaze_x", "")
    gaze_y = data.get("gaze_y", "")
    screenshot_base64 = data.get("screenshot_base64", "")

    def safe_name(text):
        return re.sub(r'[^a-zA-Z0-9_-]', '_', text)

    safe_candidate = safe_name(candidate)
    safe_reason = safe_name(reason)

    filename = f"{safe_candidate}_{safe_reason}_{uuid.uuid4().hex}.jpg"
    filepath = os.path.join(PROCTOR_SCREENSHOT_FOLDER, filename)

    if screenshot_base64.startswith("data:image"):
        img_data = screenshot_base64.split(",")[1]
        with open(filepath, "wb") as f:
            f.write(base64.b64decode(img_data))

    file_exists = os.path.isfile(PROCTOR_LOG_FILE)
    with open(PROCTOR_LOG_FILE, mode="a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if not file_exists:
            writer.writerow(["Time", "TestID", "Candidate", "Reason", "FacesDetected", "GazeX", "GazeY", "ScreenshotPath"])
        writer.writerow([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            test_id, candidate, reason, face_count,
            gaze_x, gaze_y, filepath
        ])

    print("⚠ Saved warning screenshot:", filepath)
    return jsonify({"ok": True})

@app.route("/result_webcam/<int:sid>")
def result_webcam(sid):
    return render_template("result.html", candidate="Candidate", score=None, total=None, exam_type="webcam")
# ================= VOICE INTERVIEW AUDIO UPLOAD =================
@app.route("/upload_audio/<test_id>", methods=["POST"])
@csrf.exempt
def upload_audio(test_id):
    print("✅ upload_audio route HIT")

    if not session.get("verified"):
        return jsonify({"ok": False, "error": "Unauthorized"}), 403

    candidate = session.get("candidate")
    qindex = request.form.get("question_index")
    f = request.files.get("audio")

    if not f or qindex is None:
        return jsonify({"ok": False, "error": "Audio or question index missing"}), 400

    filename = f"{uuid.uuid4().hex}.wav"
    path = os.path.join(AUDIO_FOLDER, filename)
    f.save(path)

    

    return jsonify({
        "ok": True,
        "question_index": qindex,
        "audio_file": filename,
        "transcript": ""
    })
@app.route("/evaluate/<int:submission_id>")
def evaluate_submission(submission_id):
    if not session.get("admin"):
        return redirect("/login")

    with get_db() as con:
        submission = con.execute(
            "SELECT * FROM submissions WHERE id=?",
            (submission_id,)
        ).fetchone()


        test = con.execute(
            "SELECT * FROM tests WHERE id=?",
            (submission["test_id"],)
        ).fetchone()
        # ✅ ADD THIS BLOCK
        webcam_rows = con.execute("""
            SELECT *
            FROM webcam_answers
            WHERE submission_id=?
            ORDER BY question_index
        """, (submission_id,)).fetchall()

    questions = json.loads(test["questions"])
    audio_map = {}


    answers = {}

    # Voice interview
    if test["exam_type"] == "voice":
        answers = json.loads(submission["transcript"] or "{}")

    # Webcam interview
    elif test["exam_type"] == "webcam":
        for row in webcam_rows:
            answers[str(row["question_index"])] = row["transcript"]

    # Other exams
    else:
        answers = json.loads(submission["answers"] or "{}")

    # ✅ ADD THIS
    feedback, _ = ai_evaluate(questions, answers)

    return render_template(
        "evaluate.html",
        submission=submission,
        questions=questions,
        answers=answers,
        audio_map=audio_map,
        feedback=feedback,
        webcam_rows=webcam_rows
    )

@app.route("/evaluate/save/<int:submission_id>", methods=["POST"])
def save_evaluation(submission_id):
    if not session.get("admin"):
        return redirect("/login")

    with get_db() as con:
        submission = con.execute(
            "SELECT * FROM submissions WHERE id=?",
            (submission_id,)
        ).fetchone()

        if not submission:
            flash("Submission not found", "danger")
            return redirect("/")

        test = con.execute(
            "SELECT * FROM tests WHERE id=?",
            (submission["test_id"],)
        ).fetchone()

    questions = json.loads(test["questions"])

    total_score = 0
    for i in range(len(questions)):
        total_score += int(request.form.get(str(i), 0))

    with get_db() as con:
        con.execute("""
            UPDATE submissions
            SET score=?, evaluated=1
            WHERE id=?
        """, (total_score, submission_id))

    flash("Evaluation saved successfully", "success")
    return redirect("/")

@app.route("/home")
def home():
    return render_template("home.html")

@app.route("/admin/download/<int:sid>")
def download_single_submission(sid):

    if not session.get("admin"):
        return redirect("/login")

    fmt = request.args.get("format", "csv")  # csv or xlsx

    with get_db() as con:
        r = con.execute("""
            SELECT s.*, t.title AS exam_title, t.exam_type
            FROM submissions s
            JOIN tests t ON s.test_id = t.id
            WHERE s.id=?
        """, (sid,)).fetchone()

    if not r:
        return "Submission not found", 404

    if r["exam_type"] == "voice":
        answers = json.loads(r["transcript"] or "{}")
    else:
        answers = json.loads(r["answers"] or "{}")


    base_data = {
        "Candidate": r["candidate"],
        "Email": r["candidate_email"],
        "Exam": r["exam_title"],
        "Type": r["exam_type"],
        "Score": r["score"],
        "Total": r["total"],
        "Status": "Evaluated" if r["evaluated"] else "Pending",
    }

    # ---------------- CSV ----------------
    if fmt == "csv":

        filename = f"{r['candidate']}_exam.csv"
        path = os.path.join(BASE_DIR, filename)

        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)

            for k, v in base_data.items():
                writer.writerow([k, v])

            writer.writerow([])
            writer.writerow(["Answers"])

            for q, a in answers.items():
                writer.writerow([q, a])

        return send_from_directory(BASE_DIR, filename, as_attachment=True)

    # ---------------- EXCEL ----------------
    if fmt == "xlsx":

        wb = Workbook()
        ws = wb.active
        ws.title = "Submission"

        row = 1
        for k, v in base_data.items():
            ws.cell(row=row, column=1, value=k)
            ws.cell(row=row, column=2, value=v)
            row += 1

        row += 1
        ws.cell(row=row, column=1, value="Answers")
        row += 1

        for q, a in answers.items():
            ws.cell(row=row, column=1, value=q)
            ws.cell(row=row, column=2, value=a)
            row += 1

        filename = f"{r['candidate']}_exam.xlsx"
        path = os.path.join(BASE_DIR, filename)
        wb.save(path)

        return send_from_directory(BASE_DIR, filename, as_attachment=True)

    return "Invalid format", 400

@app.route("/admin/video/<filename>")
def admin_video(filename):
    if not session.get("admin"):
        return "Unauthorized", 403

    return send_from_directory(
        VIDEO_FOLDER,
        filename,
        mimetype="video/webm"
    )
# ✅ ADD THIS NEW ROUTE
@app.route("/admin/audio/<filename>")
def admin_audio(filename):
    if not session.get("admin"):
        return "Unauthorized", 403

    return send_from_directory(
        AUDIO_FOLDER,
        filename,
        mimetype="audio/wav"
    )
@app.route("/admin/download_results")
def download_all_results():

    if not session.get("admin"):
        return redirect("/login")

    with get_db() as con:
        rows = con.execute("""
            SELECT s.*, t.title AS exam_title, t.exam_type
            FROM submissions s
            JOIN tests t ON s.test_id = t.id
            ORDER BY s.id DESC
        """).fetchall()

    filename = "all_exam_results.csv"
    path = os.path.join(BASE_DIR, filename)

    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)

        writer.writerow([
            "SubmissionID",
            "Exam",
            "Type",
            "Candidate",
            "Email",
            "Score",
            "Total",
            "Evaluated"
        ])

        for r in rows:
            writer.writerow([
                r["id"],
                r["exam_title"],
                r["exam_type"],
                r["candidate"],
                r["candidate_email"],
                r["score"],
                r["total"],
                "Yes" if r["evaluated"] else "No"
            ])

    return send_from_directory(BASE_DIR, filename, as_attachment=True)

# ---------------- RUN ----------------
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
